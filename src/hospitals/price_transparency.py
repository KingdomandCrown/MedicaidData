"""Parse CMS Hospital Price Transparency machine-readable files (MRF).

Since 2024, hospitals must publish a "standard charges" file in the CMS schema
(currently version 3.x). Two physical layouts of the same schema exist in the
wild, and this module reads both:

* **tall**  — one row per item x payer x plan. ``payer_name`` / ``plan_name``
  are their own columns. (e.g. Johns Hopkins)
* **wide**  — one row per item, with each payer/plan spread across its own set
  of columns like ``standard_charge|AETNA [1003]|AETNA PPO [100307]|negotiated_dollar``.
  These are unpivoted back into item x payer x plan rows. (e.g. U-Michigan)

Every conforming MRF begins with two metadata rows (hospital name, address,
NPI, license, version, last-updated) followed by the data header and then the
data. Some hospitals publish the data header on line 1 with no metadata
preamble at all, so the header block is *located* rather than assumed. Parsing
is driven entirely by header names, because column *order* varies between
hospitals.

The file itself may be a plain ``.csv`` or a ``.zip`` containing one CSV; both
are streamed, never fully extracted, so a 340 MB file stays memory- and
disk-friendly.
"""

from __future__ import annotations

import csv
import datetime as dt
import gzip
import io
import os
import re
import sys
import zipfile
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Iterator, Sequence

from .logging_config import get_logger
from .normalize import clean_str, parse_date

log = get_logger(__name__)

# Allow the very wide header rows / long fields these files can contain.
csv.field_size_limit(min(sys.maxsize, 2**31 - 1))

# Per-payer subfields that hang off ``standard_charge|payer|plan|<subfield>``.
_PAYER_CHARGE_SUBFIELDS = {
    "negotiated_dollar",
    "negotiated_percentage",
    "negotiated_algorithm",
    "methodology",
}
# Per-payer metric columns of the form ``<metric>|payer|plan``.
_PAYER_METRIC_FIELDS = {
    "median_amount",
    "10th_percentile",
    "90th_percentile",
    "count",
    "additional_payer_notes",
}


@dataclass
class MrfMetadata:
    """Hospital-level metadata parsed from the first two rows of an MRF."""

    hospital_name: str | None
    location_name: str | None
    hospital_address: str | None
    version: str | None
    last_updated_on: dt.date | None
    npis: list[str] = field(default_factory=list)
    license_number: str | None = None
    license_state: str | None = None
    ein: str | None = None
    source_file: str | None = None
    layout: str | None = None  # "tall" | "wide"

    @property
    def primary_npi(self) -> str | None:
        return self.npis[0] if self.npis else None


@dataclass
class ChargeRow:
    """A single normalized standard-charge fact (item x payer x plan)."""

    description: str | None
    code: str | None
    code_type: str | None
    additional_codes: str | None  # "type:code;type:code" for code|2..code|N
    billing_class: str | None
    setting: str | None
    drug_unit_of_measurement: str | None
    drug_type_of_measurement: str | None
    modifiers: str | None
    gross_charge: Decimal | None
    discounted_cash: Decimal | None
    min_charge: Decimal | None
    max_charge: Decimal | None
    payer_name: str | None
    plan_name: str | None
    negotiated_dollar: Decimal | None
    negotiated_percentage: Decimal | None
    negotiated_algorithm: str | None
    methodology: str | None
    median_amount: Decimal | None
    percentile_10: Decimal | None
    percentile_90: Decimal | None
    count: int | None
    additional_notes: str | None


# --- value parsing --------------------------------------------------------


def to_decimal(value) -> Decimal | None:
    text = clean_str(value)
    if text is None:
        return None
    text = text.replace("$", "").replace(",", "").strip()
    if not text or text.upper() in {"N/A", "NA", "NONE"}:
        return None
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def to_int(value) -> int | None:
    text = clean_str(value)
    if text is None:
        return None
    text = text.replace(",", "")
    try:
        return int(float(text))
    except (ValueError, TypeError):
        return None


# --- filename / EIN -------------------------------------------------------

_EIN_RE = re.compile(r"(\d{2})-?(\d{7})")


def ein_from_filename(path: str) -> str | None:
    """Extract the 9-digit EIN from a CMS-convention MRF filename.

    CMS names files ``<ein>_<hospital-name>_standardcharges.<ext>``. The EIN may
    be written with a dash (``24-0795959``) or without (``386006309``). Uploads
    sometimes prepend a hex hash (``<hash>-<ein>_...``), which we strip first.
    """

    base = os.path.basename(path)
    base = re.sub(r"^[0-9a-f]{6,}-", "", base, count=1)  # drop upload hash
    match = _EIN_RE.match(base) or _EIN_RE.search(base)
    if match:
        return match.group(1) + match.group(2)
    fallback = re.search(r"\d{9}", base)
    return fallback.group(0) if fallback else None


def _without_gz(path: str) -> str:
    """Path with a trailing ``.gz`` removed, for extension tests."""

    return path[:-3] if path.lower().endswith(".gz") else path


def _strip_hash_prefix(name: str) -> str:
    return re.sub(r"^[0-9a-f]{6,}-", "", os.path.basename(name), count=1)


# --- file opening ---------------------------------------------------------


# Many hospitals export these files from Excel on Windows, so they arrive in
# cp1252 rather than UTF-8 — a curly apostrophe lands as 0x92 and blows up a
# strict UTF-8 read. Sniff a leading chunk and pick the encoding that decodes
# it, preferring UTF-8 so genuinely-UTF-8 files are never mangled.
_SNIFF_BYTES = 256 * 1024


def detect_encoding(sample: bytes) -> str:
    """Return the text encoding to use for a CSV sample."""

    if sample.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    try:
        # Trailing bytes may split a multi-byte character; that is not evidence
        # against UTF-8, so ignore errors at the tail only.
        sample.decode("utf-8")
        return "utf-8-sig"
    except UnicodeDecodeError as exc:
        if exc.start >= len(sample) - 4:
            return "utf-8-sig"
    return "cp1252"


def _sniff_file(path: str) -> str:
    opener = gzip.open if path.lower().endswith(".gz") else open
    with opener(path, "rb") as fh:
        return detect_encoding(fh.read(_SNIFF_BYTES))


def open_mrf_text(path: str):
    """Return a text stream for an MRF ``.csv`` or ``.zip`` (context manager).

    For a zip, the first ``.csv`` member is streamed without extraction. The
    encoding is detected per file rather than assumed.
    """

    if path.lower().endswith(".zip"):
        return _ZipCsvStream(path)
    encoding = _sniff_file(path)
    if encoding != "utf-8-sig":
        log.info("%s: reading as %s", os.path.basename(path), encoding)
    # A .gz is decompressed on the fly, so a gzipped 4 GB export never lands
    # on disk uncompressed.
    opener = gzip.open if path.lower().endswith(".gz") else open
    return opener(path, "rt", encoding=encoding, newline="", errors="replace")


class _ZipCsvStream:
    """Context manager streaming the first CSV inside a zip as text."""

    def __init__(self, path: str):
        self._zip = zipfile.ZipFile(path)
        members = [n for n in self._zip.namelist() if n.lower().endswith(".csv")]
        if not members:
            self._zip.close()
            raise ValueError(f"no CSV member found inside {path}")
        self._member = members[0]
        with self._zip.open(self._member) as probe:
            encoding = detect_encoding(probe.read(_SNIFF_BYTES))
        if encoding != "utf-8-sig":
            log.info("%s (%s): reading as %s", os.path.basename(path), self._member, encoding)
        self._raw = self._zip.open(self._member)
        self._text = io.TextIOWrapper(
            self._raw, encoding=encoding, newline="", errors="replace"
        )

    def __enter__(self):
        return self._text

    def __exit__(self, *exc):
        self._text.close()
        self._raw.close()
        self._zip.close()


# --- metadata + layout ----------------------------------------------------


def _index_map(header: Sequence[str]) -> dict[str, int]:
    return {h.strip().lower(): i for i, h in enumerate(header)}


def parse_metadata(meta_header: Sequence[str], meta_values: Sequence[str]) -> MrfMetadata:
    idx = _index_map(meta_header)

    def val(name: str) -> str | None:
        i = idx.get(name)
        if i is None or i >= len(meta_values):
            return None
        return clean_str(meta_values[i])

    npi_raw = val("type_2_npi") or ""
    npis = [n for n in re.split(r"[|,/;\s]+", npi_raw) if n.strip().isdigit()]

    # License column is named ``license_number|<STATE>``. The value itself is
    # sometimes quoted and may embed the state again (e.g. '"1060000024|MI"').
    license_number = None
    license_state = None
    for header_name, i in idx.items():
        if header_name.startswith("license_number"):
            if "|" in header_name:
                license_state = header_name.split("|", 1)[1].upper()
            raw = clean_str(meta_values[i]) if i < len(meta_values) else None
            if raw:
                raw = raw.strip().strip('"').strip()
                if "|" in raw:
                    number, embedded_state = raw.split("|", 1)
                    raw = number.strip()
                    license_state = license_state or embedded_state.strip().upper()
            license_number = raw or None
            break

    return MrfMetadata(
        hospital_name=val("hospital_name"),
        location_name=val("location_name"),
        hospital_address=val("hospital_address"),
        version=val("version"),
        last_updated_on=parse_date(val("last_updated_on")),
        npis=npis,
        license_number=license_number,
        license_state=license_state,
    )


# Columns that only ever appear in a *data* header, never in the metadata
# preamble. Used to find the header row instead of assuming it is line 3.
_DATA_HEADER_SIGNALS = {
    "description",
    "code|1",
    "payer_name",
    "billing_class",
    "setting",
}

# How far to look for the data header before giving up. Real preambles are two
# rows plus the odd blank spacer; anything deeper is not an MRF we understand.
_MAX_PREAMBLE_ROWS = 8


def _looks_like_data_header(row: Sequence[str]) -> bool:
    lowered = {c.strip().lower() for c in row}
    if lowered & _DATA_HEADER_SIGNALS:
        return True
    return any(c.startswith("standard_charge|") for c in lowered)


def _read_header_block(reader, path: str) -> tuple[list[str], list[str], list[str]]:
    """Locate the data header and return ``(meta_header, meta_values, data_header)``.

    The CMS template puts two metadata rows above the data header, but a
    handful of hospitals publish the data header on line 1. Scanning for the
    header — rather than skipping a fixed two rows — reads both, and stops a
    header-only file from being silently parsed as zero rows.
    """

    preamble: list[list[str]] = []
    data_header: list[str] | None = None

    for row in reader:
        if not any(clean_str(c) for c in row):
            continue  # blank spacer line between the preamble and the data
        if _looks_like_data_header(row):
            data_header = list(row)
            break
        preamble.append(list(row))
        if len(preamble) > _MAX_PREAMBLE_ROWS:
            # Name the columns we did see — otherwise diagnosing an unfamiliar
            # layout means opening the file by hand.
            seen = ", ".join(c.strip() for c in preamble[0][:8] if c.strip())
            raise ValueError(
                f"{path}: no recognizable data header in the first "
                f"{_MAX_PREAMBLE_ROWS} rows (first row starts: {seen})"
            )

    if data_header is None:
        raise ValueError(f"{path}: file ended before the data header")

    if len(preamble) >= 2:
        return preamble[-2], preamble[-1], data_header
    # Header-only file: no hospital metadata to read, so the EIN from the
    # filename is the only identifier we get.
    return [], [], data_header


def detect_layout(data_header: Sequence[str]) -> str:
    """Return "tall" or "wide" from the data header column names."""

    lowered = [c.strip().lower() for c in data_header]
    if "payer_name" in lowered:
        return "tall"
    # Wide files have per-payer columns with 4 pipe-separated segments.
    for col in lowered:
        parts = col.split("|")
        if parts[0] == "standard_charge" and len(parts) == 4:
            return "wide"
    # Default to tall (single-payer / item-only files).
    return "tall"


# --- payer column grammar (wide) -----------------------------------------


def _parse_payer_column(col: str):
    """Parse a wide per-payer column into ``(payer, plan, field)`` or None."""

    parts = col.split("|")
    if parts[0] == "standard_charge" and len(parts) == 4:
        _sc, payer, plan, subfield = parts
        if subfield in _PAYER_CHARGE_SUBFIELDS:
            return payer, plan, subfield
        return None
    if parts[0] in _PAYER_METRIC_FIELDS and len(parts) == 3:
        metric, payer, plan = parts
        return payer, plan, metric
    return None


def _group_payer_columns(header: Sequence[str]) -> dict[tuple[str, str], dict[str, int]]:
    groups: dict[tuple[str, str], dict[str, int]] = {}
    for i, col in enumerate(header):
        parsed = _parse_payer_column(col.strip())
        if parsed:
            payer, plan, fieldname = parsed
            groups.setdefault((payer, plan), {})[fieldname] = i
    return groups


# --- item-level extraction ------------------------------------------------


def _collect_codes(row: Sequence[str], idx: dict[str, int]) -> tuple[str | None, str | None, str | None]:
    """Return (primary_code, primary_type, additional 'type:code;...')."""

    codes: list[tuple[str, str]] = []
    n = 1
    while f"code|{n}" in idx:
        code = clean_str(row[idx[f"code|{n}"]]) if idx[f"code|{n}"] < len(row) else None
        type_idx = idx.get(f"code|{n}|type")
        ctype = clean_str(row[type_idx]) if type_idx is not None and type_idx < len(row) else None
        if code:
            codes.append((code, ctype or ""))
        n += 1
    if not codes:
        return None, None, None
    primary_code, primary_type = codes[0]
    extra = ";".join(f"{t}:{c}" for c, t in codes[1:]) or None
    return primary_code, primary_type or None, extra


def _cell(row: Sequence[str], idx: dict[str, int], name: str):
    i = idx.get(name)
    if i is None or i >= len(row):
        return None
    return row[i]


# --- row sources ----------------------------------------------------------
#
# The tall/wide CSV logic is purely about the *rows*, so a spreadsheet is just a
# different way of producing the same list-of-cells stream. Both sources yield
# ``list[str]`` and a matching close function, and everything downstream is shared.


def _csv_row_source(path: str):
    stream_cm = open_mrf_text(path)
    stream = stream_cm.__enter__()
    return csv.reader(stream), lambda: stream_cm.__exit__(None, None, None)


def _xlsx_row_source(path: str):
    """Stream an .xlsx MRF as rows of strings (read-only, so large books stream)."""

    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - dependency is declared
        raise ValueError(
            f"{path}: reading .xlsx needs openpyxl (pip install openpyxl)"
        ) from exc

    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = book[book.sheetnames[0]]

    def rows():
        for raw in sheet.iter_rows(values_only=True):
            if raw is None:
                continue
            yield ["" if c is None else str(c) for c in raw]

    return rows(), book.close


def _row_source(path: str):
    if _without_gz(path).lower().endswith((".xlsx", ".xlsm")):
        return _xlsx_row_source(path)
    return _csv_row_source(path)


# --- top-level reader -----------------------------------------------------


def read_mrf(path: str, limit: int | None = None) -> tuple[MrfMetadata, Iterator[ChargeRow]]:
    """Open an MRF and return ``(metadata, charge_row_iterator)``.

    Reads ``.csv``, ``.zip`` (CSV inside), and ``.xlsx``/``.xlsm``. ``limit``
    caps the number of *items* (data rows) read, which is handy for
    smoke-testing very large files.
    """

    reader, close = _row_source(path)

    try:
        meta_header, meta_values, data_header = _read_header_block(reader, path)
    except Exception:
        close()
        raise

    metadata = parse_metadata(meta_header, meta_values)
    metadata.ein = ein_from_filename(path)
    metadata.source_file = _strip_hash_prefix(path)
    metadata.layout = detect_layout(data_header)
    if not meta_header:
        log.warning(
            "%s: no metadata preamble — hospital identified by EIN %s only",
            metadata.source_file,
            metadata.ein,
        )
    log.info(
        "MRF %s: %s (%s layout, NPI=%s, EIN=%s, version=%s)",
        metadata.source_file,
        metadata.hospital_name,
        metadata.layout,
        metadata.primary_npi,
        metadata.ein,
        metadata.version,
    )

    idx = _index_map(data_header)
    payer_groups = (
        _group_payer_columns(data_header) if metadata.layout == "wide" else {}
    )

    def _iter() -> Iterator[ChargeRow]:
        try:
            emitted = 0
            for n_items, row in enumerate(reader, start=1):
                if metadata.layout == "wide":
                    for charge in _wide_rows(row, idx, payer_groups):
                        emitted += 1
                        yield charge
                else:
                    charge = _tall_row(row, idx)
                    if charge is not None:
                        emitted += 1
                        yield charge
                if limit is not None and n_items >= limit:
                    break
            log.info("Parsed %d charge rows from %s", emitted, metadata.source_file)
        finally:
            close()

    return metadata, _iter()


def _item_fields(row: Sequence[str], idx: dict[str, int]) -> dict:
    code, code_type, extra = _collect_codes(row, idx)
    return dict(
        description=clean_str(_cell(row, idx, "description")),
        code=code,
        code_type=code_type,
        additional_codes=extra,
        billing_class=clean_str(_cell(row, idx, "billing_class")),
        setting=clean_str(_cell(row, idx, "setting")),
        drug_unit_of_measurement=clean_str(_cell(row, idx, "drug_unit_of_measurement")),
        drug_type_of_measurement=clean_str(_cell(row, idx, "drug_type_of_measurement")),
        modifiers=clean_str(_cell(row, idx, "modifiers")),
        gross_charge=to_decimal(_cell(row, idx, "standard_charge|gross")),
        discounted_cash=to_decimal(_cell(row, idx, "standard_charge|discounted_cash")),
        min_charge=to_decimal(_cell(row, idx, "standard_charge|min")),
        max_charge=to_decimal(_cell(row, idx, "standard_charge|max")),
    )


def _tall_row(row: Sequence[str], idx: dict[str, int]) -> ChargeRow | None:
    base = _item_fields(row, idx)
    if not base["description"] and not base["code"]:
        return None
    return ChargeRow(
        **base,
        payer_name=clean_str(_cell(row, idx, "payer_name")),
        plan_name=clean_str(_cell(row, idx, "plan_name")),
        negotiated_dollar=to_decimal(_cell(row, idx, "standard_charge|negotiated_dollar")),
        negotiated_percentage=to_decimal(_cell(row, idx, "standard_charge|negotiated_percentage")),
        negotiated_algorithm=clean_str(_cell(row, idx, "standard_charge|negotiated_algorithm")),
        methodology=clean_str(_cell(row, idx, "standard_charge|methodology")),
        median_amount=to_decimal(_cell(row, idx, "median_amount")),
        percentile_10=to_decimal(_cell(row, idx, "10th_percentile")),
        percentile_90=to_decimal(_cell(row, idx, "90th_percentile")),
        count=to_int(_cell(row, idx, "count")),
        additional_notes=clean_str(_cell(row, idx, "additional_generic_notes")),
    )


def _wide_rows(
    row: Sequence[str],
    idx: dict[str, int],
    payer_groups: dict[tuple[str, str], dict[str, int]],
) -> Iterator[ChargeRow]:
    base = _item_fields(row, idx)
    if not base["description"] and not base["code"]:
        return
    generic_notes = clean_str(_cell(row, idx, "additional_generic_notes"))

    emitted = False
    for (payer, plan), cols in payer_groups.items():
        def g(fieldname: str):
            i = cols.get(fieldname)
            return row[i] if i is not None and i < len(row) else None

        neg_dollar = to_decimal(g("negotiated_dollar"))
        neg_pct = to_decimal(g("negotiated_percentage"))
        neg_algo = clean_str(g("negotiated_algorithm"))
        if neg_dollar is None and neg_pct is None and neg_algo is None:
            continue  # this payer has no rate for this item
        emitted = True
        yield ChargeRow(
            **base,
            payer_name=payer,
            plan_name=plan,
            negotiated_dollar=neg_dollar,
            negotiated_percentage=neg_pct,
            negotiated_algorithm=neg_algo,
            methodology=clean_str(g("methodology")),
            median_amount=to_decimal(g("median_amount")),
            percentile_10=to_decimal(g("10th_percentile")),
            percentile_90=to_decimal(g("90th_percentile")),
            count=to_int(g("count")),
            additional_notes=clean_str(g("additional_payer_notes")) or generic_notes,
        )

    if not emitted:
        # Item has gross/cash but no negotiated payer rates: keep one baseline row.
        yield ChargeRow(
            **base,
            payer_name=None,
            plan_name=None,
            negotiated_dollar=None,
            negotiated_percentage=None,
            negotiated_algorithm=None,
            methodology=None,
            median_amount=None,
            percentile_10=None,
            percentile_90=None,
            count=None,
            additional_notes=generic_notes,
        )


# --- JSON layout (CMS v2.x/v3.x JSON schema) -------------------------------
#
# The JSON schema nests differently from the CSV: hospital metadata at the top,
# then a ``standard_charge_information`` array where each item has codes, a
# description, and a ``standard_charges`` array; each of those carries the
# item-level charges plus a ``payers_information`` array (one entry per payer /
# plan). We stream it with ijson so a multi-hundred-MB file stays bounded in
# memory. Files without an org NPI in the body are keyed by EIN (from the
# filename); linking then relies on the name+state heuristic or an EIN crosswalk.


import contextlib  # noqa: E402


class _BomStrippedBinary:
    """Binary stream with a leading UTF-8 BOM removed.

    Plenty of published JSON files are saved with a byte order mark. A JSON
    parser reading raw bytes sees those three bytes before the opening brace
    and rejects the document, so strip them before parsing.
    """

    _BOM = b"\xef\xbb\xbf"

    def __init__(self, stream):
        self._stream = stream
        head = stream.read(3)
        self._pending = b"" if head == self._BOM else head

    def read(self, size=-1):
        if not self._pending:
            return self._stream.read(size)
        pending, self._pending = self._pending, b""
        if size is None or size < 0:
            return pending + self._stream.read()
        if size <= len(pending):
            self._pending = pending[size:]
            return pending[:size]
        return pending + self._stream.read(size - len(pending))

    def close(self):
        self._stream.close()


@contextlib.contextmanager
def _open_binary(path: str):
    """Yield a binary stream for a ``.json`` file or a ``.json`` inside a zip."""

    if path.lower().endswith(".zip"):
        zf = zipfile.ZipFile(path)
        members = [n for n in zf.namelist() if n.lower().endswith(".json")]
        if not members:
            zf.close()
            raise ValueError(f"no JSON member found inside {path}")
        raw = zf.open(members[0])
        try:
            yield _BomStrippedBinary(raw)
        finally:
            raw.close()
            zf.close()
    else:
        opener = gzip.open if path.lower().endswith(".gz") else open
        fh = opener(path, "rb")
        try:
            yield _BomStrippedBinary(fh)
        finally:
            fh.close()


def _read_json_metadata(path: str) -> MrfMetadata:
    import ijson

    hospital_name = version = last_updated = location = None
    address: list[str] = []
    license_number = license_state = None
    npis: list[str] = []
    has_charges = False

    with _open_binary(path) as fh:
        for prefix, event, value in ijson.parse(fh):
            if event in ("string", "number"):
                if prefix == "hospital_name":
                    hospital_name = str(value)
                elif prefix == "version":
                    version = str(value)
                elif prefix == "last_updated_on":
                    last_updated = str(value)
                elif prefix in ("hospital_address", "hospital_address.item"):
                    address.append(str(value))
                elif prefix in ("hospital_location", "hospital_location.item") and location is None:
                    location = str(value)
                elif prefix == "license_information.license_number":
                    license_number = str(value)
                elif prefix == "license_information.state":
                    license_state = str(value).upper()
                elif prefix.endswith("npi") and str(value).strip().isdigit():
                    npis.append(str(value).strip())
            # Stop before streaming the (potentially huge) charges array.
            if prefix == "standard_charge_information" and event == "start_array":
                has_charges = True
                break

    # A .json (or a zip of them) that has neither the charges array nor a
    # hospital name is not an MRF at all — some other tool's data that happened
    # to land in the folder. Say so instead of loading it as an empty hospital.
    if not has_charges and not hospital_name:
        raise ValueError(f"{path}: not a CMS price transparency file")

    return MrfMetadata(
        hospital_name=hospital_name,
        location_name=location or hospital_name,
        hospital_address="; ".join(address) or None,
        version=version,
        last_updated_on=parse_date(last_updated),
        npis=npis,
        license_number=license_number,
        license_state=license_state,
    )


def _json_item_rows(item: dict) -> Iterator[ChargeRow]:
    codes = [
        (c.get("code"), c.get("type"))
        for c in (item.get("code_information") or [])
        if c.get("code")
    ]
    primary_code = primary_type = extra = None
    if codes:
        primary_code, primary_type = codes[0][0], codes[0][1]
        extra = ";".join(f"{t or ''}:{c}" for c, t in codes[1:]) or None

    drug = item.get("drug_information") or {}
    description = clean_str(item.get("description"))

    for sc in item.get("standard_charges") or []:
        modifiers = sc.get("modifiers")
        if isinstance(modifiers, list):
            modifiers = ";".join(str(m) for m in modifiers) or None
        base = dict(
            description=description,
            code=clean_str(primary_code),
            code_type=clean_str(primary_type),
            additional_codes=extra,
            billing_class=clean_str(sc.get("billing_class")),
            setting=clean_str(sc.get("setting")),
            drug_unit_of_measurement=clean_str(drug.get("unit")),
            drug_type_of_measurement=clean_str(drug.get("type")),
            modifiers=clean_str(modifiers),
            gross_charge=to_decimal(sc.get("gross_charge")),
            discounted_cash=to_decimal(sc.get("discounted_cash")),
            min_charge=to_decimal(sc.get("minimum")),
            max_charge=to_decimal(sc.get("maximum")),
        )
        payers = sc.get("payers_information") or []
        if payers:
            for p in payers:
                yield ChargeRow(
                    **base,
                    payer_name=clean_str(p.get("payer_name")),
                    plan_name=clean_str(p.get("plan_name")),
                    negotiated_dollar=to_decimal(p.get("standard_charge_dollar")),
                    negotiated_percentage=to_decimal(p.get("standard_charge_percentage")),
                    negotiated_algorithm=clean_str(p.get("standard_charge_algorithm")),
                    methodology=clean_str(p.get("methodology")),
                    median_amount=to_decimal(p.get("estimated_amount")),
                    percentile_10=None,
                    percentile_90=None,
                    count=None,
                    additional_notes=clean_str(p.get("additional_payer_notes")),
                )
        else:
            yield ChargeRow(
                **base,
                payer_name=None,
                plan_name=None,
                negotiated_dollar=None,
                negotiated_percentage=None,
                negotiated_algorithm=None,
                methodology=None,
                median_amount=None,
                percentile_10=None,
                percentile_90=None,
                count=None,
                additional_notes=None,
            )


def read_mrf_json(path: str, limit: int | None = None) -> tuple[MrfMetadata, Iterator[ChargeRow]]:
    """Open a JSON MRF and return ``(metadata, charge_row_iterator)``, streamed."""

    import ijson

    metadata = _read_json_metadata(path)
    metadata.ein = ein_from_filename(path)
    metadata.source_file = _strip_hash_prefix(path)
    metadata.layout = "json"
    log.info(
        "MRF %s: %s (json layout, NPI=%s, EIN=%s, version=%s)",
        metadata.source_file,
        metadata.hospital_name,
        metadata.primary_npi,
        metadata.ein,
        metadata.version,
    )

    def _iter() -> Iterator[ChargeRow]:
        emitted = 0
        with _open_binary(path) as fh:
            for n_items, item in enumerate(
                ijson.items(fh, "standard_charge_information.item"), start=1
            ):
                for charge in _json_item_rows(item):
                    emitted += 1
                    yield charge
                if limit is not None and n_items >= limit:
                    break
        log.info("Parsed %d charge rows from %s", emitted, metadata.source_file)

    return metadata, _iter()


def _is_json_source(path: str) -> bool:
    lower = _without_gz(path).lower()
    if lower.endswith(".json"):
        return True
    if lower.endswith(".zip"):
        with zipfile.ZipFile(path) as zf:
            names = [n.lower() for n in zf.namelist()]
        has_json = any(n.endswith(".json") for n in names)
        has_csv = any(n.endswith(".csv") for n in names)
        return has_json and not has_csv
    return False


def read_any(path: str, limit: int | None = None) -> tuple[MrfMetadata, Iterator[ChargeRow]]:
    """Read a price-transparency file, dispatching on JSON vs CSV/XLSX layout."""

    if _is_json_source(path):
        return read_mrf_json(path, limit=limit)
    return read_mrf(path, limit=limit)
