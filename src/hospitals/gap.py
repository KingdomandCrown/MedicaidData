"""Work out which hospitals still need a price-transparency file downloaded.

The knowledge base has two halves that do not line up on their own: the POS
``hospitals`` table is the *universe* (every Medicare-certified hospital, keyed
by CCN), and ``charge_sources`` is what has actually been ingested (keyed by
EIN/NPI). The gap between them is the download worklist.

Ordering that worklist matters more than producing it. One visit to a health
system's transparency page often yields a dozen hospitals, so systems are ranked
ahead of independents, and systems we have *already* pulled from are ranked
ahead of systems we have not found yet — for those the URL is already known and
the remaining hospitals are usually on the same page.

System membership is not in the POS file, so it is inferred from the hospital
name's leading brand token ("BAYLOR SCOTT", "ASCENSION", "CHRISTUS"). That is a
heuristic and is labelled as such in the output: it groups the obvious chains
and leaves genuinely independent hospitals alone, which is all the ranking
needs.
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from urllib.parse import quote_plus

from sqlalchemy import select
from sqlalchemy.engine import Engine

from .db import charge_sources, hospitals
from .link import _state_of, normalize_name
from .logging_config import get_logger

log = get_logger(__name__)

# Words that describe a hospital rather than name its owner. A brand made only
# of these is not a brand, so the leading token is not enough on its own.
_GENERIC = {
    "HOSPITAL", "HOSPITALS", "MEDICAL", "CENTER", "CENTRE", "HEALTH",
    "HEALTHCARE", "REGIONAL", "MEMORIAL", "COMMUNITY", "COUNTY", "GENERAL",
    "SYSTEM", "CLINIC", "CARE", "DISTRICT", "AREA", "VALLEY", "SERVICES",
    "CHILDRENS", "REHABILITATION", "BEHAVIORAL", "SURGICAL", "SPECIALTY",
    "UNIVERSITY", "CITY", "COUNTRY", "STATE",
}

# Leading words too common to identify a system by themselves — a second token
# is needed ("SAINT LUKES" and "SAINT FRANCIS" are different systems).
_NEEDS_SECOND_TOKEN = {
    "SAINT", "ST", "NORTH", "SOUTH", "EAST", "WEST", "NEW", "GOOD", "HOLY",
    "GREATER", "CENTRAL", "FIRST", "MOUNT", "LAKE", "PORT", "FORT", "SAN",
    "SANTA", "LOS", "LAS", "EL", "LA",
}


def system_key(name: str | None) -> str | None:
    """Infer a health-system brand from a hospital name.

    Returns ``None`` when the name carries no usable brand — those hospitals
    are treated as independents.
    """

    tokens = normalize_name(name).split()
    if not tokens:
        return None

    first = tokens[0]
    if first in _NEEDS_SECOND_TOKEN and len(tokens) > 1:
        return f"{first} {tokens[1]}"
    if first in _GENERIC:
        # Lead with a descriptor ("COMMUNITY HOSPITAL OF X"): take the first
        # token that actually names something.
        for token in tokens[1:]:
            if token not in _GENERIC and len(token) >= 4:
                return token
        return None
    if len(first) < 4:
        return f"{first} {tokens[1]}" if len(tokens) > 1 else None
    return first


@dataclass
class SystemGap:
    """A candidate system with hospitals still to download."""

    system: str
    remaining: int
    states: list[str]
    already_publishing: bool
    known_source: str | None
    examples: list[str] = field(default_factory=list)


@dataclass
class HospitalGap:
    ccn: str
    name: str
    city: str | None
    state: str | None
    subtype: str | None
    beds: int | None
    system: str | None


@dataclass
class StateCoverage:
    state: str
    total: int
    covered: int

    @property
    def remaining(self) -> int:
        return self.total - self.covered

    @property
    def pct(self) -> float:
        return round(100.0 * self.covered / self.total, 1) if self.total else 0.0


@dataclass
class Unattributed:
    """A file we hold whose hospital could not be identified.

    Usually a file published with no metadata preamble and no EIN in its name,
    so there is nothing to join on. These are *not* gaps — the data is already
    in hand — but they will look like gaps until someone assigns a CCN, so they
    are reported separately rather than silently counted either way.
    """

    source_file: str
    hospital_name: str | None
    ein: str | None
    primary_npi: str | None
    state: str | None
    charge_count: int


@dataclass
class GapReport:
    systems: list[SystemGap]
    independents: list[HospitalGap]
    coverage: list[StateCoverage]
    unattributed: list[Unattributed]
    total_hospitals: int
    total_covered: int

    @property
    def total_remaining(self) -> int:
        return self.total_hospitals - self.total_covered

    @property
    def uncrawled_states(self) -> list[str]:
        """States where nothing at all has been ingested yet."""

        return [c.state for c in self.coverage if c.covered == 0]


def _covered_ccns(conn) -> tuple[set[str], dict[str, str], list[Unattributed]]:
    """Return (CCNs downloaded, brand -> example source file, unattributed files).

    A charge source counts as covering a hospital when it was linked to a CCN,
    or when its name and state match a POS hospital unambiguously — the same
    heuristic the linker uses, applied here so an unlinked database still gives
    a usable answer.
    """

    name_index: dict[tuple[str, str], set[str]] = {}
    for h in conn.execute(select(hospitals.c.ccn, hospitals.c.name, hospitals.c.state)):
        if not h.ccn or not h.state:
            continue
        key = (normalize_name(h.name), h.state.upper())
        if key[0]:
            name_index.setdefault(key, set()).add(h.ccn)

    covered: set[str] = set()
    publishing: dict[str, str] = {}
    unattributed: list[Unattributed] = []

    rows = conn.execute(
        select(
            charge_sources.c.ccn,
            charge_sources.c.hospital_name,
            charge_sources.c.license_state,
            charge_sources.c.hospital_address,
            charge_sources.c.source_file,
            charge_sources.c.charge_count,
            charge_sources.c.ein,
            charge_sources.c.primary_npi,
        )
    ).mappings()

    for src in rows:
        if not src["charge_count"]:
            continue  # parsed but empty: not actually covered
        brand = system_key(src["hospital_name"])
        if brand:
            publishing.setdefault(brand, src["source_file"])
        if src["ccn"]:
            covered.add(src["ccn"])
            continue
        state = _state_of(src)
        candidates = (
            name_index.get((normalize_name(src["hospital_name"]), state))
            if state
            else None
        )
        if candidates and len(candidates) == 1:
            covered.add(next(iter(candidates)))
            continue
        unattributed.append(
            Unattributed(
                source_file=src["source_file"],
                hospital_name=src["hospital_name"],
                ein=src["ein"],
                primary_npi=src["primary_npi"],
                state=state,
                charge_count=src["charge_count"],
            )
        )

    unattributed.sort(key=lambda u: -u.charge_count)
    return covered, publishing, unattributed


def build_gap_report(engine: Engine, *, min_system_size: int = 2) -> GapReport:
    """Diff the POS hospital universe against what has been ingested."""

    with engine.connect() as conn:
        covered, publishing, unattributed = _covered_ccns(conn)

        universe = conn.execute(
            select(
                hospitals.c.ccn,
                hospitals.c.name,
                hospitals.c.city,
                hospitals.c.state,
                hospitals.c.provider_subtype,
                hospitals.c.certified_bed_count,
            ).where(hospitals.c.is_active.is_(True))
        ).mappings().all()

    per_state_total: dict[str, int] = defaultdict(int)
    per_state_covered: dict[str, int] = defaultdict(int)
    remaining: list[HospitalGap] = []

    for h in universe:
        state = (h["state"] or "??").upper()
        per_state_total[state] += 1
        if h["ccn"] in covered:
            per_state_covered[state] += 1
            continue
        remaining.append(
            HospitalGap(
                ccn=h["ccn"],
                name=h["name"],
                city=h["city"],
                state=state,
                subtype=h["provider_subtype"],
                beds=h["certified_bed_count"],
                system=system_key(h["name"]),
            )
        )

    clusters: dict[str, list[HospitalGap]] = defaultdict(list)
    for gap in remaining:
        if gap.system:
            clusters[gap.system].append(gap)

    systems: list[SystemGap] = []
    grouped: set[str] = set()
    for brand, members in clusters.items():
        if len(members) < min_system_size:
            continue
        grouped.update(m.ccn for m in members)
        systems.append(
            SystemGap(
                system=brand,
                remaining=len(members),
                states=sorted({m.state for m in members if m.state}),
                already_publishing=brand in publishing,
                known_source=publishing.get(brand),
                examples=[m.name for m in sorted(members, key=lambda m: m.name)[:3]],
            )
        )

    # Known publishers first (their URL is already in hand), then by how many
    # hospitals one visit would yield.
    systems.sort(key=lambda s: (not s.already_publishing, -s.remaining, s.system))

    independents = sorted(
        (g for g in remaining if g.ccn not in grouped),
        key=lambda g: (g.state or "", g.name),
    )

    coverage = sorted(
        (
            StateCoverage(state=st, total=total, covered=per_state_covered.get(st, 0))
            for st, total in per_state_total.items()
        ),
        key=lambda c: (c.pct, c.state),
    )

    report = GapReport(
        systems=systems,
        independents=independents,
        coverage=coverage,
        unattributed=unattributed,
        total_hospitals=len(universe),
        total_covered=len(universe) - len(remaining),
    )
    log.info(
        "Gap report: %d of %d hospitals covered; %d remaining "
        "(%d candidate systems, %d independents, %d unattributed files)",
        report.total_covered,
        report.total_hospitals,
        report.total_remaining,
        len(systems),
        len(independents),
        len(unattributed),
    )
    return report


# --- spreadsheet ----------------------------------------------------------


_SHEETS = (
    "Priority Systems",
    "Independents by State",
    "State Coverage",
    "Unattributed Files",
)


# CMS requires every hospital to publish a machine-readable index at this path
# on its own domain, which is the shortest route to the file once you know the
# domain. Neither the POS file nor Hospital Enrollments carries a website, so
# the domain is the one thing this cannot supply.
CMS_HPT_INDEX_PATH = "/cms-hpt.txt"

_SEARCH_BASE = "https://www.google.com/search?q="


def mrf_search_url(name: str, city: str | None = None, state: str | None = None) -> str:
    """A search that lands on the hospital's price-transparency page.

    The honest limit: no public CMS dataset maps a hospital to its website, so
    a real file URL cannot be derived. What removes most of the manual work is
    a query already narrowed to the right words — the quoted name pins the
    hospital, and the rest are the terms these pages actually use.
    """

    terms = [f'"{name}"']
    if city:
        terms.append(city)
    if state:
        terms.append(state)
    terms += ["price transparency", "machine readable", "standard charges"]
    return _SEARCH_BASE + quote_plus(" ".join(terms))


def write_xlsx(report: GapReport, path: str) -> str:
    """Write the report to a three-sheet workbook and return the path."""

    try:
        import openpyxl
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - dependency is declared
        raise ValueError("writing .xlsx needs openpyxl (pip install openpyxl)") from exc

    book = openpyxl.Workbook()
    bold = Font(bold=True)

    def sheet(
        title: str,
        headers: list[str],
        rows: list[list],
        first: bool = False,
        link_column: int | None = None,
    ):
        ws = book.active if first else book.create_sheet()
        ws.title = title
        ws.append(headers)
        for cell in ws[1]:
            cell.font = bold
        for row in rows:
            ws.append(row)
        if link_column is not None:
            # The URL is long and the same shape every time, so show a label
            # and hang the link off it — a column of visible query strings
            # would push everything else off the screen.
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=link_column)
                url = cell.value
                if url:
                    cell.value = "Find MRF"
                    cell.hyperlink = url
                    cell.style = "Hyperlink"
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for i, header in enumerate(headers, start=1):
            longest = max([len(header)] + [len(str(r[i - 1] or "")) for r in rows] or [0])
            ws.column_dimensions[get_column_letter(i)].width = min(60, longest + 2)
        return ws

    sheet(
        _SHEETS[0],
        [
            "Priority",
            "Candidate system",
            "Hospitals to download",
            "States",
            "Source already known",
            "Known file (same publisher)",
            "Example hospitals",
            "Find MRF",
        ],
        [
            [
                n,
                s.system.title(),
                s.remaining,
                ", ".join(s.states),
                "YES" if s.already_publishing else "",
                s.known_source or "",
                "; ".join(s.examples),
                mrf_search_url(s.system.title(), state=s.states[0] if s.states else None),
            ]
            for n, s in enumerate(report.systems, start=1)
        ],
        first=True,
        link_column=8,
    )

    sheet(
        _SHEETS[1],
        ["State", "CCN", "Hospital", "City", "Type", "Certified beds", "Find MRF"],
        [
            [
                g.state,
                g.ccn,
                g.name,
                g.city or "",
                g.subtype or "",
                g.beds or "",
                mrf_search_url(g.name, g.city, g.state),
            ]
            for g in report.independents
        ],
        link_column=7,
    )

    sheet(
        _SHEETS[2],
        ["State", "Active hospitals", "Downloaded", "Remaining", "% covered", "Status"],
        [
            [
                c.state,
                c.total,
                c.covered,
                c.remaining,
                c.pct,
                "not started" if c.covered == 0 else "",
            ]
            for c in report.coverage
        ],
    )

    sheet(
        _SHEETS[3],
        [
            "Source file",
            "Hospital name in file",
            "EIN",
            "NPI",
            "State",
            "Charge rows",
            "Assign CCN",
        ],
        [
            [
                u.source_file,
                u.hospital_name or "",
                u.ein or "",
                u.primary_npi or "",
                u.state or "",
                u.charge_count,
                "",
            ]
            for u in report.unattributed
        ],
    )

    notes = book.create_sheet()
    notes.title = "How to find a file"
    for line in [
        ["Finding a hospital's machine-readable file"],
        [],
        ["1.", "Click 'Find MRF' on either worklist sheet. The search is already"],
        ["", "narrowed to the hospital and the words these pages use."],
        [],
        ["2.", "Once you know the hospital's domain, skip the search entirely:"],
        ["", f"https://<their-domain>{CMS_HPT_INDEX_PATH}"],
        ["", "CMS requires every hospital to publish that index, and it names the"],
        ["", "file's real URL. It is the fastest route and the one a crawler"],
        ["", "should try first."],
        [],
        ["3.", "Work the Priority Systems sheet before the independents: one"],
        ["", "system page often carries every facility it owns, so a single"],
        ["", "download can cover dozens of hospitals."],
        [],
        ["Note", "No public CMS dataset maps a hospital to its website, so this"],
        ["", "workbook cannot give you the file URL directly. The search link is"],
        ["", "the closest thing the available data supports."],
    ]:
        notes.append(line)
    notes.column_dimensions["A"].width = 8
    notes.column_dimensions["B"].width = 78
    notes["A1"].font = bold

    book.save(path)
    log.info("Wrote %s", path)
    return path
