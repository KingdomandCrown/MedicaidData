#!/usr/bin/env python3
"""
xlsx_to_resolver_input.py

Convert a per-state hospital workbook (one sheet per state, AHD-style
"Table of Search Results" export) into the input CSV expected by
scripts/resolve_hospital_websites.py.

Each sheet carries a preamble of search-criteria rows above the real table,
and the header row's position varies by sheet, so the header is located by
its first cell ("Hospital Name") rather than by row number. Expected table
columns (extra columns are ignored):

    Hospital Name | CMS Certification Number | Beds | City | State | ZIP | Telephone | ...

Output columns match the resolver's defaults exactly:

    ccn, hospital_name, address, city, state, zip

The source export has no street address, so `address` is left blank — the
resolver builds its Places query from name + city + state + zip and treats
the address as optional.

Rows without a CCN (satellite campuses, sub-units billed under a parent
facility) cannot be keyed into the knowledge base and are written to a
separate `*_no_ccn.csv` sidecar for manual review instead of being silently
dropped. Duplicate CCNs across sheets keep the first occurrence.

Requires: pip install openpyxl
"""

import argparse
import csv
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: this script needs openpyxl (pip install openpyxl).")

HEADER_SENTINEL = "hospital name"

# Header-cell text -> output field. Matched case-insensitively after
# whitespace collapsing, so minor formatting drift in the export is tolerated.
COLUMN_MAP = {
    "hospital name": "hospital_name",
    "cms certification number": "ccn",
    "city": "city",
    "state": "state",
    "zip": "zip",
}

OUTPUT_FIELDS = ["ccn", "hospital_name", "address", "city", "state", "zip"]


def clean(value) -> str:
    """Collapse whitespace and stringify; empty/None become ''."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_ccn(value) -> str:
    """Match the knowledge base's CCN rules: 6 chars, zero-pad short numerics,
    upper-case alphanumerics (federal/military CCNs like 02013F)."""
    text = clean(value).replace(" ", "").upper()
    if text.isdigit() and len(text) < 6:
        text = text.zfill(6)
    return text


def normalize_zip(value) -> str:
    """5-digit ZIP; Excel stores these as ints, dropping leading zeros."""
    digits = re.sub(r"\D", "", clean(value))
    if not digits:
        return ""
    return digits[:5].zfill(5) if len(digits) >= 5 else digits.zfill(5)


def find_header(rows) -> tuple[int, dict] | None:
    """Locate the header row and map output fields to column indexes."""
    for i, row in enumerate(rows):
        first = clean(row[0]).lower() if row else ""
        if first == HEADER_SENTINEL:
            indexes = {}
            for col, cell in enumerate(row):
                key = clean(cell).lower()
                if key in COLUMN_MAP:
                    indexes[COLUMN_MAP[key]] = col
            return i, indexes
    return None


def extract_sheet(ws) -> tuple[list[dict], int]:
    """Return (records, skipped_blank) for one worksheet."""
    rows = list(ws.iter_rows(values_only=True))
    located = find_header(rows)
    if located is None:
        return [], 0
    header_idx, cols = located

    records = []
    for row in rows[header_idx + 1 :]:
        name = clean(row[cols["hospital_name"]]) if "hospital_name" in cols else ""
        if not name:
            continue  # blank spacer/footer rows
        records.append(
            {
                "ccn": normalize_ccn(row[cols["ccn"]]) if "ccn" in cols else "",
                "hospital_name": name,
                "address": "",
                "city": clean(row[cols["city"]]) if "city" in cols else "",
                "state": clean(row[cols["state"]]).upper() if "state" in cols else "",
                "zip": normalize_zip(row[cols["zip"]]) if "zip" in cols else "",
            }
        )
    return records, 0


def convert(input_path: str, output_path: str, no_ccn_path: str) -> dict:
    wb = openpyxl.load_workbook(input_path, read_only=True)

    keyed: dict[str, dict] = {}
    no_ccn: list[dict] = []
    duplicates = 0
    sheets_missing_table = []

    for sheet in wb.sheetnames:
        records, _ = extract_sheet(wb[sheet])
        if not records:
            sheets_missing_table.append(sheet)
            continue
        for rec in records:
            if not rec["ccn"]:
                no_ccn.append(rec)
            elif rec["ccn"] in keyed:
                duplicates += 1
            else:
                keyed[rec["ccn"]] = rec

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_FIELDS)
        writer.writeheader()
        for rec in sorted(keyed.values(), key=lambda r: r["ccn"]):
            writer.writerow(rec)

    with open(no_ccn_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_FIELDS)
        writer.writeheader()
        for rec in sorted(no_ccn, key=lambda r: (r["state"], r["hospital_name"])):
            writer.writerow(rec)

    return {
        "sheets": len(wb.sheetnames),
        "sheets_missing_table": sheets_missing_table,
        "with_ccn": len(keyed),
        "no_ccn": len(no_ccn),
        "duplicates": duplicates,
    }


def main():
    parser = argparse.ArgumentParser(
        description="Convert a per-state hospital XLSX into resolver input CSV."
    )
    parser.add_argument("--input", required=True, help="Path to the .xlsx workbook")
    parser.add_argument(
        "--output",
        default="hospital_input.csv",
        help="Resolver input CSV to write (default: hospital_input.csv)",
    )
    parser.add_argument(
        "--no-ccn-output",
        default=None,
        help="Sidecar CSV for rows without a CCN (default: <output>_no_ccn.csv)",
    )
    args = parser.parse_args()

    no_ccn_path = args.no_ccn_output or re.sub(r"\.csv$", "", args.output) + "_no_ccn.csv"
    stats = convert(args.input, args.output, no_ccn_path)

    print(f"Sheets read: {stats['sheets']}")
    if stats["sheets_missing_table"]:
        print(f"  WARNING - no table found in: {', '.join(stats['sheets_missing_table'])}")
    print(f"Hospitals with CCN (resolver input): {stats['with_ccn']} -> {args.output}")
    print(f"Rows without CCN (manual review):    {stats['no_ccn']} -> {no_ccn_path}")
    print(f"Duplicate CCNs skipped:              {stats['duplicates']}")


if __name__ == "__main__":
    main()
