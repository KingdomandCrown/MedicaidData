#!/usr/bin/env python3
"""
extract_ahd_websites.py

Harvest hospital websites from American Hospital Directory (AHD) profile
exports (.xlsx, one hospital per file).

CONTEXT
-------
AHD profile reports are mostly CMS-derived (HCRIS cost-report financials,
Medicare claims utilization, Hospital Compare quality) — but no CMS public
use file carries a website field. The "Hospital Website" / "System Website"
rows in an AHD profile are AHD's own editorial enrichment, which makes them
a high-quality, CCN-keyed website source in their own right.

This script walks a folder of AHD profile exports, locates the labeled rows
on each profile sheet (positions drift between exports, so rows are found
by label, not index), and writes one CSV row per hospital:

    ccn, hospital_name, website_uri, system_website, operating_status, source

The output is directly usable as a --results input to
verify_hospital_websites.py (it has `ccn`, `website_uri`, and `source`
columns), where AHD counts as an independent source for cross-source
agreement alongside Places / Wikidata / OSM.

Check that your AHD subscription terms permit this use of exported reports.

USAGE
-----
    pip install openpyxl
    python scripts/extract_ahd_websites.py --input-dir ahd_reports/ \
        --output hospital_websites_ahd.csv
"""

import argparse
import csv
import glob
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: this script needs openpyxl (pip install openpyxl).")

# Profile-sheet labels -> output fields. Compared case-insensitively after
# whitespace collapsing.
LABELS = {
    "cms certification number": "ccn",
    "hospital website": "website_uri",
    "system website": "system_website",
    "operating status": "operating_status",
}

OUTPUT_FIELDS = [
    "ccn", "hospital_name", "website_uri", "system_website",
    "operating_status", "source", "file",
]


def clean(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).replace("\xa0", " ").strip()


def normalize_ccn(value) -> str:
    text = clean(value).replace(" ", "").upper()
    if text.isdigit() and len(text) < 6:
        text = text.zfill(6)
    return text


def normalize_website(url: str) -> str:
    url = clean(url)
    if url and not re.match(r"^https?://", url, re.IGNORECASE):
        url = "https://" + url
    return url


def extract_profile(ws) -> dict:
    """Pull labeled fields from an AHD profile sheet."""
    out = {v: "" for v in LABELS.values()}
    name = ""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        cells = [clean(c) for c in row]
        first = cells[0].lower() if cells and cells[0] else ""
        # The hospital name is the first standalone line after the
        # "Profile - <date>" banner.
        if not name and i > 0 and cells and cells[0] and not first.startswith("profile"):
            name = cells[0]
        if first in LABELS and len(cells) > 1:
            field = LABELS[first]
            if not out[field]:
                out[field] = cells[1]
        # Fallback: banner form "CMS Certification Number: 171376".
        if not out["ccn"]:
            m = re.match(r"cms certification number:\s*(\S+)", first)
            if m:
                out["ccn"] = m.group(1)
    out["hospital_name"] = name
    out["ccn"] = normalize_ccn(out["ccn"])
    out["website_uri"] = normalize_website(out["website_uri"])
    out["system_website"] = normalize_website(out["system_website"])
    return out


def extract_file(path: str) -> dict | None:
    wb = openpyxl.load_workbook(path, read_only=True)
    # The profile is the first sheet; the rest are Departments/Financial/etc.
    profile = extract_profile(wb[wb.sheetnames[0]])
    if not profile["ccn"]:
        return None
    profile["source"] = "ahd"
    profile["file"] = os.path.basename(path)
    return profile


def main():
    parser = argparse.ArgumentParser(description="Extract websites from AHD profile exports.")
    parser.add_argument("--input-dir", required=True, help="Folder containing AHD .xlsx exports")
    parser.add_argument("--output", default="hospital_websites_ahd.csv")
    args = parser.parse_args()

    paths = sorted(glob.glob(os.path.join(args.input_dir, "*.xlsx")))
    if not paths:
        sys.exit(f"No .xlsx files found in {args.input_dir}")

    extracted, no_ccn, no_site = 0, [], []
    with open(args.output, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_FIELDS)
        writer.writeheader()
        for path in paths:
            row = extract_file(path)
            if row is None:
                no_ccn.append(os.path.basename(path))
                continue
            if not row["website_uri"]:
                no_site.append(os.path.basename(path))
            writer.writerow(row)
            extracted += 1

    print(f"Extracted {extracted}/{len(paths)} profiles -> {args.output}")
    if no_ccn:
        print(f"  WARNING - no CCN found (skipped): {', '.join(no_ccn)}")
    if no_site:
        print(f"  NOTE - profile has no Hospital Website row: {', '.join(no_site)}")


if __name__ == "__main__":
    main()
